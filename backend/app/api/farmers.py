from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from app.core.database import get_db
from app.core.security import get_current_user
from app.models.user import User
from app.models.farmer import Farmer
from app.models.purchase import Purchase
from app.schemas.farmer import FarmerCreate, FarmerUpdate, FarmerResponse
from app.schemas.common import ApiResponse
from app.services.farmer_service import FarmerService
from pydantic import BaseModel
from typing import Optional
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime

router = APIRouter(prefix="/api/farmers", tags=["农户"])


class ExportRequest(BaseModel):
    farmer_ids: Optional[list[int]] = None


@router.get("")
async def list_farmers(
    keyword: str | None = None,
    offset: int = 0,
    limit: int = 200,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    service = FarmerService(db)
    if keyword:
        farmers = await service.search(keyword, offset=offset, limit=min(limit, 500))
    else:
        farmers = await service.get_all(offset=offset, limit=min(limit, 500))
    return ApiResponse(data=[FarmerResponse.model_validate(f) for f in farmers])


@router.post("")
async def create_farmer(
    req: FarmerCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    service = FarmerService(db)
    farmer = await service.get_or_create(req.name)
    if req.phone or req.id_card:
        farmer = await service.update(
            farmer_id=farmer.id,
            phone=req.phone,
            id_card=req.id_card,
        )
    return ApiResponse(data=FarmerResponse.model_validate(farmer))


@router.post("/export")
async def export_farmers(
    req: ExportRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if req.farmer_ids:
        result = await db.execute(select(Farmer).where(Farmer.id.in_(req.farmer_ids), Farmer.deleted == "0"))
    else:
        result = await db.execute(select(Farmer).where(Farmer.deleted == "0"))
    farmers = list(result.scalars().all())

    farmer_id_map = {f.id: f for f in farmers}
    result = await db.execute(
        select(Purchase).where(Purchase.farmer_id.in_(farmer_id_map.keys()), Purchase.deleted == False)
        .order_by(Purchase.created_at)
    )
    purchases = list(result.scalars().all())

    wb = Workbook()
    ws = wb.active
    ws.title = "农户收购记录"

    bold = Font(bold=True, size=12)
    header_font = Font(bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")

    # Summary
    ws.append(["共计"])
    ws["A1"].font = bold
    total_gross = sum(float(p.gross_weight or 0) for p in purchases)
    total_empty = sum(float(p.empty_weight or 0) for p in purchases)
    total_net = sum(float(p.net_weight or 0) for p in purchases)
    total_amount = sum(float(p.total_amount or 0) for p in purchases)
    ws.append([f"农户数: {len(farmers)}, 收购单数: {len(purchases)}"])
    ws.append([f"总毛重: {total_gross:.2f}, 总皮重: {total_empty:.2f}, 总净重: {total_net:.2f}, 总金额: {total_amount:.2f}"])
    ws.append([])

    # Detail header
    headers = ["农户姓名", "身份证号", "电话", "粮食品种", "毛重", "皮重", "净重", "单价", "总金额", "创建时间", "空车称重时间", "结账时间"]
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = header_font
        cell.alignment = center

    for p in purchases:
        farmer = farmer_id_map.get(p.farmer_id)
        ws.append([
            farmer.name if farmer else "",
            farmer.id_card if farmer else "",
            farmer.phone if farmer else "",
            p.grain_type,
            float(p.gross_weight or 0),
            float(p.empty_weight or 0),
            float(p.net_weight or 0),
            float(p.unit_price or 0),
            float(p.total_amount or 0),
            p.created_at.strftime("%Y-%m-%d %H:%M:%S") if p.created_at else "",
            p.empty_weighted_at.strftime("%Y-%m-%d %H:%M:%S") if p.empty_weighted_at else "",
            p.completed_at.strftime("%Y-%m-%d %H:%M:%S") if p.completed_at else "",
        ])

    # Column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 16

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"farmers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/{farmer_id}")
async def get_farmer(
    farmer_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    service = FarmerService(db)
    farmer = await service.get_by_id(farmer_id)
    if not farmer or farmer.deleted == "1":
        raise HTTPException(status_code=404, detail="农户不存在")
    return ApiResponse(data=FarmerResponse.model_validate(farmer))


@router.put("/{farmer_id}")
async def update_farmer(
    farmer_id: int,
    req: FarmerUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    service = FarmerService(db)
    try:
        farmer = await service.update(
            farmer_id=farmer_id,
            name=req.name,
            phone=req.phone,
            id_card=req.id_card,
        )
        return ApiResponse(data=FarmerResponse.model_validate(farmer))
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
